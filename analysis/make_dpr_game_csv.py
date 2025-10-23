#!/usr/bin/env python3
"""make_dpr_game_csv.py

Generate a tidy CSV (and optionally an .xlsx grid) with the *DPR* empirical
payoff game for a given holding-period value.

The output format is identical to `dpr_game_hp140.csv` but all missing payoff
cells are **imputed** so that every (own-MELO-count, other-MELO-count) pair has a
numerical value.

Imputation strategy (choose with `--impute`):
    mean   – replace each NaN with the *mean payoff of that strategy*.
    rowcol – arithmetic mean of the non-NaN values in the same row & column
              (falls back to the overall strategy mean if still NaN).

Usage examples
--------------
    # Replicate the HP 140 table with simple mean imputation
    python analysis/make_dpr_game_csv.py --hp 140 --out analysis/dpr_game_hp140_full.csv

    # Build HP 100 table with smarter neighbour imputation and also an XLSX
    python analysis/make_dpr_game_csv.py --hp 100 --impute neighbor \
                                         --xlsx analysis/dpr_game_hp100_full.xlsx
"""
from __future__ import annotations

import argparse, csv, itertools, math, sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook

# ---------------------------------------------------------------------------
# We re-use helpers defined in analysis/plot_bootstrap_compare.py to avoid code
# duplication.  (Importing that file is cheap and does not execute the plotting
# logic when we only need the helper functions.)
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from analysis.plot_bootstrap_compare import (
    collect_profiles_by_seed,
    build_game_from_profiles,
)

CANONICAL_STRATS = [
    ("MOBI", "M-ELO", "MOBI_0_100_shade0_0"),
    ("MOBI", "CDA",   "MOBI_100_0_shade250_500"),
    ("ZI",   "M-ELO", "ZI_0_100_shade250_500"),
    ("ZI",   "CDA",   "ZI_100_0_shade250_500"),
]

# ---------------------------------------------------------------------------
# Small helper to fill NaNs in a 5×5 grid
# ---------------------------------------------------------------------------

def _impute_grid(grid: np.ndarray, *, mode: str = "mean") -> np.ndarray:
    """Return a copy of *grid* with NaNs replaced according to *mode*."""
    out = grid.copy()
    if mode == "mean":
        mu = np.nanmean(out)
        out[np.isnan(out)] = mu
        return out

    if mode == "neighbor":
        filled = grid.copy()
        # iterative neighbour averaging until no NaNs resolved further
        while True:
            updated = False
            nan_positions = np.argwhere(np.isnan(filled))
            if nan_positions.size == 0:
                break
            for i, j in nan_positions:
                neigh = []
                for di, dj in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
                    ii, jj = i + di, j + dj
                    if 0 <= ii < 5 and 0 <= jj < 5 and not math.isnan(filled[ii, jj]):
                        neigh.append(filled[ii, jj])
                if neigh:
                    filled[i, j] = float(np.mean(neigh))
                    updated = True
            if not updated:
                break  # nothing new could be filled
        # fall back to row/col mean then global
        return _impute_grid(filled, mode="rowcol")

    if mode == "rowcol":
        for i in range(5):
            for j in range(5):
                if not math.isnan(out[i, j]):
                    continue
                row_vals = out[i, :]
                col_vals = out[:, j]
                vals = np.concatenate([row_vals[~np.isnan(row_vals)], col_vals[~np.isnan(col_vals)]])
                if vals.size:
                    out[i, j] = np.mean(vals)
        # Fall back to global mean if any NaNs remain
        if np.isnan(out).any():
            mu = np.nanmean(out)
            out[np.isnan(out)] = mu
        return out

    raise ValueError(f"unknown imputation mode: {mode}")

# ---------------------------------------------------------------------------
# Main routine
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--hp", type=int, required=True, help="Holding-period value (e.g. 140)")
    parser.add_argument("--out", type=Path, required=True, help="Output CSV path")
    parser.add_argument("--impute", choices=["mean", "rowcol", "neighbor"], default="neighbor",
                        help="Imputation strategy for missing payoffs")
    parser.add_argument("--xlsx", type=Path, help="Optional XLSX output path (grids)")
    parser.add_argument("--template-csv", type=Path, help="Optional CSV using the specialised DPR table layout")
    # NEW: identical layout but written to XLSX
    parser.add_argument("--template-xlsx", type=Path, help="Optional XLSX using the specialised DPR table layout")
    args = parser.parse_args()

    # --------------------------------------------------
    # 1.  Build pooled empirical game for this HP
    # --------------------------------------------------
    profiles = list(itertools.chain.from_iterable(collect_profiles_by_seed(args.hp).values()))
    if not profiles:
        raise SystemExit(f"No payoff profiles found for HP = {args.hp}.")

    game = build_game_from_profiles(profiles)

    # --------------------------------------------------
    # 2.  Extract *per-agent* payoffs (no further scaling)  — DPR definition
    # --------------------------------------------------
    # Payoff table layout: 4 strategies × 25 columns (0…4 MELO vs 0…4 MELO)
    table = game.game.rsg_payoff_table.clone().numpy()

    # Treat exact zeros as missing (set to NaN) – often indicates unobserved profile
    table[table == 0.0] = np.nan

    # --------------------------------------------------
    # 3.  Emit tidy rows with OPTIONAL imputation
    # --------------------------------------------------
    rows: list[dict[str, object]] = []
    for s_idx, (role, nice_name, _) in enumerate(CANONICAL_STRATS):
        grid = table[s_idx].reshape(5, 5)
        grid = _impute_grid(grid, mode=args.impute)
        for own in range(5):
            for other in range(5):
                rows.append({
                    "role": role,
                    "strategy": nice_name,
                    "own_M-ELO_count": own,
                    "other_M-ELO_count": other,
                    "payoff": float(grid[own, other]),
                })

    # --------------------------------------------------
    # 4.  Save CSV  (resolve() avoids relative-path issues on print)
    # --------------------------------------------------
    df = pd.DataFrame(rows)
    out_path = args.out.resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(out_path, index=False, quoting=csv.QUOTE_MINIMAL)
    try:
        rel = out_path.relative_to(Path.cwd())
    except ValueError:
        rel = out_path  # fallback to absolute
    print(f"Saved {rel}  ({len(df)} rows)")

    # --------------------------------------------------
    # 5.  Optional XLSX grid dump
    # --------------------------------------------------
    if args.xlsx:
        wb = Workbook()
        for s_idx, (role, nice_name, _) in enumerate(CANONICAL_STRATS):
            ws = wb.create_sheet(f"{role}_{nice_name}")
            grid = table[s_idx].reshape(5, 5)
            grid = _impute_grid(grid, mode=args.impute)
            # header row / col
            ws.cell(row=1, column=1, value="own\\other")
            for k in range(5):
                ws.cell(row=1, column=2 + k, value=k)
                ws.cell(row=2 + k, column=1, value=k)
            for i in range(5):
                for j in range(5):
                    ws.cell(row=2 + i, column=2 + j, value=float(grid[i, j]))
        # remove default sheet
        if "Sheet" in wb.sheetnames:
            std = wb["Sheet"]
            wb.remove(std)
        x_path = args.xlsx.resolve()
        x_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(x_path)
        try:
            relx = x_path.relative_to(Path.cwd())
        except ValueError:
            relx = x_path
        print(f"Saved {relx}")

    # --------------------------------------------------
    # 6.  Optional template-style CSV / XLSX (advisor’s format)
    # --------------------------------------------------
    if args.template_csv or args.template_xlsx:
        tmpl_rows = []
        # Header rows -----------------------------------------------------
        tmpl_rows.append([
            "Bkg trader payoff", "", "other Bkg", "MOBI", "", "MOBI trader payoff", "", "other MOBI", "Bkg",
        ])
        tmpl_rows.append([
            "M-ELO", "Lit", "#M-ELO", "#M-ELO", "", "M-ELO", "Lit", "#M-ELO", "#M-ELO",
        ])

        # Helper to fetch payoff (handling NaN → empty string)
        def payoff(role, strat_idx, own, other):
            # Mapping: (MOBI,M-ELO)=0, (MOBI,CDA)=1, (ZI,M-ELO)=2, (ZI,CDA)=3
            base = 0 if role == "MOBI" else 2
            idx = base + strat_idx  # strat_idx: 0=M-ELO, 1=CDA/Lit
            val = grid_cache[idx][own, other]
            return "" if np.isnan(val) else float(val)

        # Build a small grid cache (after imputation) to avoid recompute
        grid_cache = []
        for s_idx, _ in enumerate(CANONICAL_STRATS):
            g = table[s_idx].reshape(5, 5)
            g = _impute_grid(g, mode=args.impute)
            grid_cache.append(g)

        # Fill 25 rows (own=0..4, other=0..4)
        for own in range(5):
            for other in range(5):
                tmpl_rows.append([
                    payoff("MOBI", 0, own, other),   # Bkg M-ELO
                    payoff("MOBI", 1, own, other),   # Bkg Lit/CDA
                    own, other,                       # counts for Bkg block
                    "",                               # blank separator column
                    payoff("ZI", 0, own, other),     # MOBI M-ELO (ZI role)
                    payoff("ZI", 1, own, other),     # MOBI Lit/CDA (ZI role)
                    own, other,                       # counts for MOBI block
                ])

        # --- 6a. CSV -----------------------------------------------------
        if args.template_csv:
            t_path = args.template_csv.resolve()
            t_path.parent.mkdir(parents=True, exist_ok=True)
            with t_path.open("w", newline="") as fh:
                csv.writer(fh).writerows(tmpl_rows)
            relt = t_path
            try:
                relt = t_path.relative_to(Path.cwd())
            except ValueError:
                pass
            print(f"Saved {relt}")

        # --- 6b. XLSX ----------------------------------------------------
        if args.template_xlsx:
            wb = Workbook()
            ws = wb.active
            ws.title = "Payoffs"
            for r, row in enumerate(tmpl_rows, 1):
                for c, value in enumerate(row, 1):
                    ws.cell(row=r, column=c, value=value)

            # widen first and sixth columns a bit for readability
            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["F"].width = 18

            x_path = args.template_xlsx.resolve()
            x_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(x_path)
            relx = x_path
            try:
                relx = x_path.relative_to(Path.cwd())
            except ValueError:
                pass
            print(f"Saved {relx}")

if __name__ == "__main__":
    main() 